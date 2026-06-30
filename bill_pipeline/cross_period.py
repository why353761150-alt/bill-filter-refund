"""
跨账期退款分离 + 校验 + 大额支出标黄。

输入：data/working/<账期>/中国银行_清理结果.xlsx
      （人工需在"剔除退款后账单"sheet 的"备注"列填入：
        "上期还款" / "上期账单退款，已抵扣上期账单还款" / "上期账单退款，已抵扣本期账单还款"）

输出：data/output/<账期>/中国银行_最终账单.xlsx
      （精简：3 个 sheet —— 最终账单 / 跨账期退款 / 汇总）
"""
from __future__ import annotations

from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

from .config import (
    TRANSACTION_COLUMNS,
    SUMMARY_COLUMNS,
    CROSS_PERIOD_NOTES,
    CLEAN_RESULT_SUFFIX,
    FINAL_BILL_NAME,
    HIGHLIGHT_THRESHOLD,
    log,
    period_dirs,
    identify_period,
    current_period,
)
from .excel_utils import (
    parse_money,
    normalize_money_column,
    ensure_transaction_columns,
    ensure_summary_columns,
    drop_total_rows,
    add_total_row,
    apply_header_format,
    apply_body_border,
    auto_column_width,
    apply_total_row_format,
    highlight_high_value_expenses,
    _resolve_column,
)


def split_cross_period(clean_result_path: Path | str) -> Path:
    """
    从清理结果中分离跨账期退款记录，校验上下期欠款，
    输出最终账单（精简 3 sheet）。
    """
    clean_path = Path(clean_result_path)
    if not clean_path.exists():
        raise FileNotFoundError(f"找不到清理结果: {clean_path}")

    period = identify_period(clean_path.name) or current_period()
    dirs = period_dirs(period)
    output_path = dirs["output"] / FINAL_BILL_NAME

    log.info(f"🔀 分离跨账期退款: {clean_path.name}")

    # 读取原始账单（用于校验）和清理后账单
    original_df = pd.read_excel(clean_path, sheet_name="原始账单")
    original_df = ensure_transaction_columns(original_df)
    original_df["存入"] = normalize_money_column(original_df["存入"])
    original_df["支出"] = normalize_money_column(original_df["支出"])
    # 摘要列：用 _resolve_column 兼容旧列名
    for target_col in SUMMARY_COLUMNS:
        actual = _resolve_column(pd.read_excel(clean_path, sheet_name="原始账单"), target_col)
        if actual and actual in original_df.columns:
            original_df[target_col] = pd.to_numeric(original_df[actual], errors="coerce")
        elif target_col in original_df.columns:
            original_df[target_col] = pd.to_numeric(original_df[target_col], errors="coerce")

    cleaned_df = pd.read_excel(clean_path, sheet_name="剔除退款后账单")
    cleaned_df = ensure_transaction_columns(cleaned_df)
    cleaned_df["存入"] = normalize_money_column(cleaned_df["存入"])
    cleaned_df["支出"] = normalize_money_column(cleaned_df["支出"])

    # 移除合计行后筛选
    cleaned_no_total = drop_total_rows(cleaned_df)
    cleaned_no_total["备注"] = cleaned_no_total["备注"].fillna("").astype(str).str.strip()

    cross_period_mask = cleaned_no_total["备注"] != ""
    cross_period_df = cleaned_no_total[cross_period_mask].copy()
    final_cleaned_df = cleaned_no_total[~cross_period_mask].reset_index(drop=True)

    log.info(f"   跨账期记录: {len(cross_period_df)} 条")
    log.info(f"   最终账单记录: {len(final_cleaned_df)} 条")

    # ========== 校验：上期欠款 ==========
    last_repayment = cross_period_df.loc[
        cross_period_df["备注"].str.contains(CROSS_PERIOD_NOTES["LAST_REPAYMENT"], na=False),
        "存入",
    ].sum()

    last_refund_offset_last = _sum_with_sign(
        cross_period_df[cross_period_df["备注"].str.contains(
            CROSS_PERIOD_NOTES["LAST_REFUND_OFFSET_LAST"], na=False
        )]
    )
    theoretical_last = last_repayment + last_refund_offset_last

    last_col = "上期欠款余额"  # 已经 ensure_summary_columns 重命名
    actual_last = pd.to_numeric(
        original_df.loc[~original_df["交易日"].astype(str).str.contains("合计", na=False),
                       last_col].dropna(), errors="coerce"
    ).iloc[0] if not original_df.empty and last_col in original_df.columns and original_df[last_col].notna().any() else 0
    diff_last = theoretical_last - float(actual_last or 0)

    # ========== 校验：本期欠款 ==========
    final_expenditure = final_cleaned_df["支出"].sum()
    final_deposit = final_cleaned_df["存入"].sum()
    this_refund_offset = _sum_with_sign(
        cross_period_df[cross_period_df["备注"].str.contains(
            CROSS_PERIOD_NOTES["LAST_REFUND_OFFSET_THIS"], na=False
        )]
    )
    theoretical_this = final_expenditure - this_refund_offset
    # 月度应还（用户在最终账单合计行下方看到的明确数字）
    monthly_payment = final_expenditure - final_deposit
    this_col = "本期欠款余额"
    actual_this = pd.to_numeric(
        original_df.loc[~original_df["交易日"].astype(str).str.contains("合计", na=False),
                       this_col].dropna(), errors="coerce"
    ).iloc[0] if not original_df.empty and this_col in original_df.columns and original_df[this_col].notna().any() else 0
    diff_this = theoretical_this - float(actual_this or 0)

    # ========== 构造最终输出 ==========
    # 1. 最终账单：剔除跨账期退款后的 7 列 + 合计行
    final_with_total = add_total_row(final_cleaned_df[TRANSACTION_COLUMNS])

    # 2. 跨账期退款明细（按备注分组排序）
    cross_period_out = cross_period_df[TRANSACTION_COLUMNS].copy()
    if not cross_period_out.empty:
        cross_period_out = cross_period_out.sort_values(["备注", "交易日"])

    # 3. 汇总
    summary_data = {
        "项目": [
            "上期还款", "上期账单退款(抵上期)", "理论上期欠款",
            "实际上期欠款", "上期校验差值",
            "最终支出", "上期账单退款(抵本期)", "理论本期欠款",
            "实际本期欠款", "本期校验差值",
            "原始记录数", "清理后记录数", "跨账期记录数",
        ],
        "金额": [
            last_repayment, last_refund_offset_last, theoretical_last,
            float(actual_last or 0), diff_last,
            final_expenditure, this_refund_offset, theoretical_this,
            float(actual_this or 0), diff_this,
            len(original_df), len(cleaned_no_total), len(cross_period_df),
        ],
    }
    summary_df = pd.DataFrame(summary_data)

    # ========== 写 Excel ==========
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        final_with_total.to_excel(writer, sheet_name="最终账单", index=False)
        cross_period_out.to_excel(writer, sheet_name="跨账期退款", index=False)
        summary_df.to_excel(writer, sheet_name="汇总", index=False)

    # 格式美化
    wb = load_workbook(output_path)
    for name in ["最终账单", "跨账期退款", "汇总"]:
        ws = wb[name]
        apply_header_format(ws)
        apply_body_border(ws)
        auto_column_width(ws)

    # 给"最终账单"加合计行格式 + 大额标黄 + 本月应还行
    if "最终账单" in wb.sheetnames:
        ws = wb["最终账单"]
        if ws.max_row >= 2 and ws.cell(row=ws.max_row, column=1).value == "合计":
            apply_total_row_format(ws)

            # ========== 新增"本月应还"行 ==========
            from openpyxl.styles import Font, PatternFill
            from .excel_utils import YELLOW_FILL, HEADER_FILL, THIN_BORDER, TOTAL_FILL

            payment_row = ws.max_row + 1
            ws.cell(row=payment_row, column=1, value="本月应还").font = Font(bold=True, color="FFFFFF")
            ws.cell(row=payment_row, column=1).fill = PatternFill(
                start_color="C00000", end_color="C00000", fill_type="solid"
            )
            ws.cell(row=payment_row, column=1).border = THIN_BORDER

            # 公式：=F合计 - E合计（F=支出，E=存入）
            total_row = payment_row - 1
            ws.cell(row=payment_row, column=5, value=f"=F{total_row}-E{total_row}").font = Font(bold=True, color="FFFFFF")
            ws.cell(row=payment_row, column=5).fill = PatternFill(
                start_color="C00000", end_color="C00000", fill_type="solid"
            )
            ws.cell(row=payment_row, column=5).border = THIN_BORDER
            # 其余格子留空但也填色
            for col in [2, 3, 4, 6, 7]:
                ws.cell(row=payment_row, column=col).fill = PatternFill(
                    start_color="C00000", end_color="C00000", fill_type="solid"
                )
                ws.cell(row=payment_row, column=col).border = THIN_BORDER
            log.info(f"   已在'最终账单'sheet 添加'本月应还'行（红底白字）")

        # 大额支出标黄（"支出"列在 TRANSACTION_COLUMNS 中索引为 5，列字母 F）
        from .config import TRANSACTION_COLUMNS as TC
        expense_col_idx = TC.index("支出") + 1
        highlight_high_value_expenses(
            ws, expense_col=expense_col_idx, threshold=HIGHLIGHT_THRESHOLD
        )
        log.info(f"   已对 >{HIGHLIGHT_THRESHOLD} 元的支出标黄")

    # ========== 校验：把本期差值标红写入"汇总"sheet ==========
    if "汇总" in wb.sheetnames and abs(diff_this) > 0.01:
        ws = wb["汇总"]
        from openpyxl.styles import Font as F2
        # "本期校验差值" 在 summary_data 中的索引 = 9（第 10 行）
        # 找它的实际行号（表头在第 1 行）
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == "本期校验差值":
                ws.cell(row=row, column=2).font = F2(bold=True, color="C00000")
                log.warning(f"   ⚠️ 本期校验差值 {diff_this:.2f} ≠ 0，已在'汇总'sheet 标红")
                break

    wb.save(output_path)

    # ========== 打印校验结果 ==========
    log.info("=" * 50)
    log.info("📊 校验结果")
    log.info("=" * 50)
    log.info(f"上期欠款: 理论={theoretical_last:.2f}, 实际={float(actual_last or 0):.2f}, "
             f"差值={diff_last:.2f} {'✓' if abs(diff_last) < 0.01 else '✗'}")
    log.info(f"本期欠款: 理论={theoretical_this:.2f}, 实际={float(actual_this or 0):.2f}, "
             f"差值={diff_this:.2f} {'✓' if abs(diff_this) < 0.01 else '✗'}")
    log.info(f"本月应还（支出合计 - 存入合计）= {monthly_payment:.2f}")
    log.info(f"✅ 最终账单: {output_path}")
    return output_path


def _sum_with_sign(df: pd.DataFrame) -> float:
    """存入记为正、支出记为负，求和"""
    if df.empty:
        return 0.0
    return float(df["存入"].sum() - df["支出"].sum())
