"""
通用 Excel 工具：
- 金额清洗
- 合计行公式
- 黄色填充（标黄指定行）
- 高额支出自动标黄
- 标准化的 DataFrame 写入
"""
from __future__ import annotations

import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .config import TRANSACTION_COLUMNS, log


# ============================================================
# 单元格格式预设
# ============================================================
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
HEADER_FILL = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

HEADER_FONT = Font(bold=True, size=11)
BODY_FONT = Font(size=11)
TOTAL_FONT = Font(bold=True, size=11)


# ============================================================
# 金额处理
# ============================================================
_MONEY_CLEAN = re.compile(r"[^\d.\-]")


def parse_money(value) -> float:
    """鲁棒的金额解析：处理 str/float/None/含逗号/含¥/含公式"""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s or s == "nan":
        return 0.0
    # 处理公式（如 =SUM(...)），先尝试 eval
    if s.startswith("="):
        try:
            expr = s[1:].lstrip("=")
            # 防止危险：仅允许数字/运算符/括号/字母函数
            if re.match(r"^[A-Z0-9().,+\-*/\s]+$", expr, re.IGNORECASE):
                return float(eval(expr))
        except Exception:
            return 0.0
    # 清洗字符串
    s = _MONEY_CLEAN.sub("", s)
    try:
        return float(s) if s else 0.0
    except (ValueError, TypeError):
        return 0.0


def normalize_money_column(series: pd.Series) -> pd.Series:
    """将整个 Series 转为 float，非法值填 0"""
    return series.apply(parse_money)


# ============================================================
# DataFrame 标准化
# ============================================================
# 兼容新旧两种列名（中英混合长名 → 纯中文短名）
# 例如 "交易日 Transaction Date" → "交易日"
LEGACY_COLUMN_ALIASES = {
    "交易日": ["交易日", "交易日 Transaction Date", "交易日期"],
    "银行记账日": ["银行记账日", "银行记账日 Posting Date", "记账日期"],
    "卡号后四位": ["卡号后四位", "卡号后四位 Last Four Digits of Card Number", "卡号"],
    "交易描述": ["交易描述", "交易描述 Description", "摘要", "交易摘要"],
    "存入": ["存入", "存入 Deposit", "Deposit", "收入"],
    "支出": ["支出", "支出 Expenditure", "Expenditure", "消费"],
    "备注": ["备注", "Notes", "说明"],
    # 摘要列
    "上期欠款余额": ["上期欠款余额", "上期欠款", "上期余额"],
    "本期支出金额": ["本期支出金额", "本期支出"],
    "本期存入金额": ["本期存入金额", "本期存入"],
    "本期欠款余额": ["本期欠款余额", "本期欠款", "本期余额"],
}


def _resolve_column(df: pd.DataFrame, target: str) -> str | None:
    """从 df 的列中找匹配 target 别名的真实列名，找不到返回 None"""
    aliases = LEGACY_COLUMN_ALIASES.get(target, [target])
    for col in df.columns:
        if str(col).strip() in aliases:
            return col
    return None


def ensure_transaction_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    标准化交易明细 DataFrame 为 7 列结构。
    兼容旧版本（中英混合长名列名）：自动重命名到标准短列名。
    """
    df = df.copy()
    rename_map = {}
    for target in TRANSACTION_COLUMNS:
        actual = _resolve_column(df, target)
        if actual and actual != target:
            rename_map[actual] = target
    if rename_map:
        df = df.rename(columns=rename_map)
    # 补全缺失列
    for col in TRANSACTION_COLUMNS:
        if col not in df.columns:
            df[col] = ""
    return df[TRANSACTION_COLUMNS]


def ensure_summary_columns(df: pd.DataFrame) -> pd.DataFrame:
    """标准化摘要 DataFrame 为 4 列结构（兼容旧版本）"""
    df = df.copy()
    rename_map = {}
    for target in SUMMARY_COLUMNS:
        actual = _resolve_column(df, target)
        if actual and actual != target:
            rename_map[actual] = target
    if rename_map:
        df = df.rename(columns=rename_map)
    for col in SUMMARY_COLUMNS:
        if col not in df.columns:
            df[col] = None
    return df[SUMMARY_COLUMNS]


def drop_total_rows(df: pd.DataFrame, date_col: str = "交易日") -> pd.DataFrame:
    """移除合计行（"合计"/"总计"/"total"）"""
    if date_col not in df.columns:
        return df
    mask = df[date_col].astype(str).str.contains("合计|总计|total", case=False, na=False)
    return df[~mask].copy()


def add_total_row(df: pd.DataFrame, deposit_col: str = "存入",
                  expense_col: str = "支出") -> pd.DataFrame:
    """
    在 df 末尾添加合计行，金额列使用 Excel 公式。
    要求 df 已用 TRANSACTION_COLUMNS 标准化。
    """
    df = drop_total_rows(df)
    if df.empty:
        return df

    data_row_count = len(df)
    # Excel 公式：第二行到第 data_row_count+1 行
    sum_range_dep = f"E2:E{data_row_count + 1}"
    sum_range_exp = f"F2:F{data_row_count + 1}"

    total_row = {col: "" for col in TRANSACTION_COLUMNS}
    total_row["交易日"] = "合计"
    total_row["存入"] = f"=SUM({sum_range_dep})"
    total_row["支出"] = f"=SUM({sum_range_exp})"

    total_df = pd.DataFrame([total_row], columns=TRANSACTION_COLUMNS)
    return pd.concat([df, total_df], ignore_index=True)


# ============================================================
# 写入后的格式调整
# ============================================================
def apply_total_row_format(ws, sheet_name: str | None = None):
    """
    给工作表最后一行（合计行）应用格式。
    调用方需先确认最后一行是合计行。
    """
    if ws.max_row < 2:
        return
    last_row = ws.max_row
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=last_row, column=col)
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER


def apply_header_format(ws, header_row: int = 1):
    """表头格式化"""
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def apply_body_border(ws, start_row: int = 2):
    """给数据区加边框"""
    for row in range(start_row, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER


def auto_column_width(ws, widths: dict[str, int] | None = None):
    """设置列宽（按列字母），未指定的列根据内容自动估算"""
    for col_letter, width in (widths or {}).items():
        ws.column_dimensions[col_letter].width = width
    # 自动估算未指定的列
    for col_cells in ws.columns:
        col_letter = col_cells[0].column_letter
        if col_letter in (widths or {}):
            continue
        max_len = 0
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 50)


# ============================================================
# 标黄工具
# ============================================================
def highlight_rows(ws, row_indices: list[int], columns: int | None = None,
                   start_row_offset: int = 2):
    """
    把指定行整行标黄。
    row_indices: 0-based 的 DataFrame 行索引；写入 Excel 时 +start_row_offset
    columns: 标黄的列数（None = 整行）
    """
    n_cols = columns or ws.max_column
    for idx in row_indices:
        excel_row = idx + start_row_offset
        for col in range(1, n_cols + 1):
            ws.cell(row=excel_row, column=col).fill = YELLOW_FILL


def highlight_high_value_expenses(ws, expense_col: int = 6,
                                  threshold: float = 200,
                                  header_row: int = 1):
    """
    把支出金额 > threshold 的行整行标黄。
    expense_col: "支出"列在 TRANSACTION_COLUMNS 中的位置（默认 6，即 F 列）。
    """
    n_cols = ws.max_column
    for row in range(header_row + 1, ws.max_row + 1):
        cell = ws.cell(row=row, column=expense_col)
        if cell.value is None or cell.value == "":
            continue
        if str(cell.value).startswith("="):  # 跳过合计行公式
            continue
        amount = parse_money(cell.value)
        if amount > threshold:
            for col in range(1, n_cols + 1):
                ws.cell(row=row, column=col).fill = YELLOW_FILL


# ============================================================
# 多 sheet 写入的统一封装
# ============================================================
def write_workbook(output_path, sheets: dict[str, pd.DataFrame],
                   post_process: callable = None):
    """
    写入多 sheet Excel，并对每个 sheet 应用统一格式。
    post_process(ws, sheet_name, df) 用于追加特殊格式（图表等）。
    """
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for name, df in sheets.items():
            df.to_excel(writer, sheet_name=name, index=False)

        if post_process:
            wb = writer.book
            for name in sheets:
                ws = wb[name]
                post_process(ws, name, sheets[name])


def beautify_worksheet(ws, df: pd.DataFrame, total_row: bool = True,
                      highlight_threshold: float | None = None,
                      expense_col: int = 6):
    """
    一站式美化：表头 + 边框 + 列宽 + 合计行 + 可选大额标黄。
    """
    apply_header_format(ws)
    apply_body_border(ws)
    auto_column_width(ws)

    if total_row and not df.empty:
        apply_total_row_format(ws)

    if highlight_threshold is not None:
        highlight_high_value_expenses(ws, expense_col=expense_col,
                                       threshold=highlight_threshold)
