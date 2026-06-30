"""
退款候选生成 + 人工标注后的账单清理。

两步：
1. match_refunds(bill_path) → 生成退款候选清单（含黄色标注需人工筛选的行）
2. apply_refunds(bill_path, refund_list_path) → 应用人工标注，生成清理结果
"""
from __future__ import annotations

from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

from .config import (
    TRANSACTION_COLUMNS,
    SUMMARY_COLUMNS,
    RAW_BILL_NAME,
    REFUND_CANDIDATE_SUFFIX,
    CLEAN_RESULT_SUFFIX,
    TEMPLATE_AMOUNT,
    HIGHLIGHT_THRESHOLD,
    log,
    period_dirs,
    identify_period,
    current_period,
)
from .excel_utils import (
    parse_money,
    normalize_money_column,
    drop_total_rows,
    add_total_row,
    ensure_transaction_columns,
    apply_total_row_format,
    apply_header_format,
    apply_body_border,
    auto_column_width,
    highlight_rows,
    highlight_high_value_expenses,
    YELLOW_FILL,
)


# ============================================================
# 步骤 1：生成退款候选清单
# ============================================================
def match_refunds(bill_path: Path | str) -> Path:
    """
    从原始账单中找出所有金额匹配的"存入-支出"对，输出候选清单。
    一笔存入对应多笔支出的记录标黄（需人工筛选）。
    """
    bill_path = Path(bill_path)
    if not bill_path.exists():
        raise FileNotFoundError(f"找不到账单: {bill_path}")

    period = identify_period(bill_path.name) or current_period()
    dirs = period_dirs(period)
    output_path = dirs["working"] / bill_path.name.replace(".xlsx", REFUND_CANDIDATE_SUFFIX)

    log.info(f"🔍 读取账单: {bill_path}")
    df = pd.read_excel(bill_path, sheet_name="原始账单")
    df = ensure_transaction_columns(df)

    # 转数值
    df["存入"] = normalize_money_column(df["存入"])
    df["支出"] = normalize_money_column(df["支出"])

    # 拆分存入 / 支出
    deposits = df[df["存入"] > 0].sort_values("交易日").reset_index()
    expenses = df[df["支出"] > 0].sort_values("交易日").reset_index()

    # 为每笔存入找所有金额匹配且日期 ≤ 存入日期的支出
    candidates = []
    for _, dep in deposits.iterrows():
        matched = expenses[
            (abs(expenses["支出"] - dep["存入"]) < TEMPLATE_AMOUNT)
            & (expenses["交易日"] <= dep["交易日"])
        ]
        for _, exp in matched.iterrows():
            candidates.append({
                "存入原索引": int(dep["index"]),
                "存入交易日": dep["交易日"],
                "存入描述": dep["交易描述"],
                "存入金额": dep["存入"],
                "支出原索引": int(exp["index"]),
                "支出交易日": exp["交易日"],
                "支出描述": exp["交易描述"],
                "支出金额": exp["支出"],
            })

    candidates_df = pd.DataFrame(candidates)
    if candidates_df.empty:
        log.info("💡 未发现金额匹配的退款候选")
        candidates_df = pd.DataFrame(columns=[
            "序号", "存入交易日", "存入描述", "存入金额",
            "支出交易日", "支出描述", "支出金额",
        ])
    else:
        # 按存入日期排序
        candidates_df = candidates_df.sort_values(
            ["存入交易日", "支出交易日"]
        ).reset_index(drop=True)
        candidates_df.insert(0, "序号", range(1, len(candidates_df) + 1))

    # 识别需人工筛选的组（一笔存入对应多笔支出）
    need_manual_set: set[int] = set()
    if not candidates_df.empty and "存入交易日" in candidates_df.columns:
        groups = candidates_df.groupby(["存入交易日", "存入金额"])
        for (d, a), g in groups:
            if len(g) > 1:
                need_manual_set.update(g.index.tolist())

    # 写 Excel
    summary_df = pd.DataFrame({
        "项目": [
            "候选记录总数", "需人工筛选记录数",
            "涉及的存入笔数", "自动可判定笔数",
        ],
        "数值": [
            len(candidates_df),
            len(need_manual_set),
            candidates_df.groupby(["存入交易日", "存入金额"]).ngroups if not candidates_df.empty else 0,
            len(candidates_df) - len(need_manual_set),
        ],
    })

    instructions = pd.DataFrame({
        "步骤": [
            "1. 本表已按存入时间排序，存入在左、支出在右",
            "2. 黄色填充的行表示一笔存入对应多笔支出（需人工筛选）",
            "3. 请删除不是真正退款的行（一笔存入最多保留一条支出）",
            "4. 删除多余的合计行（最后一行的'合计'）",
            "5. 将本表 sheet 重命名为'本期退款账单'",
            "6. 保存后执行 clean 步骤",
        ]
    })

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # 原始账单（保留以便对照）—— 摘要列可能缺失（PDF 解析失败时），做安全切片
        base_cols = TRANSACTION_COLUMNS + [c for c in SUMMARY_COLUMNS if c in df.columns]
        df[base_cols].to_excel(writer, sheet_name="原始账单", index=False)
        # candidates_df 可能为空（无退款候选），但必须写出表头
        if candidates_df.empty:
            pd.DataFrame(columns=[
                "序号", "存入交易日", "存入描述", "存入金额",
                "支出交易日", "支出描述", "支出金额",
            ]).to_excel(writer, sheet_name="退款候选", index=False)
        else:
            candidates_df.to_excel(writer, sheet_name="退款候选", index=False)
        summary_df.to_excel(writer, sheet_name="统计", index=False)
        instructions.to_excel(writer, sheet_name="使用说明", index=False)

    # 后处理：标黄 + 格式
    wb = load_workbook(output_path)
    for sheet_name in ["退款候选", "统计", "使用说明", "原始账单"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        apply_header_format(ws)
        apply_body_border(ws)
        auto_column_width(ws)

    if "退款候选" in wb.sheetnames and need_manual_set:
        highlight_rows(wb["退款候选"], list(need_manual_set))

    wb.save(output_path)
    log.info(f"✅ 退款候选清单: {output_path}")
    log.info(f"   候选 {len(candidates_df)} 条，其中 {len(need_manual_set)} 条需人工筛选")
    return output_path


# ============================================================
# 步骤 2：应用人工标注的退款清单 → 清理账单
# ============================================================
def _read_refund_list(refund_path: Path) -> pd.DataFrame:
    """读取人工标注过的退款清单，兼容多种工作表名"""
    xls = pd.ExcelFile(refund_path)
    sheet = None
    for name in ["本期退款账单", "退款候选"]:
        if name in xls.sheet_names:
            sheet = name
            break
    if sheet is None:
        raise ValueError(
            f"退款清单中找不到 '本期退款账单' 或 '退款候选' 工作表: {xls.sheet_names}"
        )
    df = pd.read_excel(xls, sheet_name=sheet)
    # 移除合计行
    for col in ["序号", "存入交易日", "支出交易日"]:
        if col in df.columns:
            df = df[~df[col].astype(str).str.contains(
                "合计|总计|total", case=False, na=False
            )]
    return df


def apply_refunds(bill_path: Path | str,
                  refund_list_path: Path | str) -> Path:
    """
    根据人工标注的退款清单，剔除原始账单中的退款记录。
    输出的清理结果包含三个 sheet：原始账单 / 本期退款账单 / 剔除退款后账单。
    """
    bill_path = Path(bill_path)
    refund_path = Path(refund_list_path)

    if not bill_path.exists():
        raise FileNotFoundError(f"找不到账单: {bill_path}")
    if not refund_path.exists():
        raise FileNotFoundError(f"找不到退款清单: {refund_path}")

    period = identify_period(bill_path.name) or current_period()
    dirs = period_dirs(period)
    output_path = dirs["working"] / bill_path.name.replace(".xlsx", CLEAN_RESULT_SUFFIX)

    log.info(f"🧹 清理账单: {bill_path}")

    original_df = pd.read_excel(bill_path, sheet_name="原始账单")
    original_df = ensure_transaction_columns(original_df)
    original_df["存入"] = normalize_money_column(original_df["存入"])
    original_df["支出"] = normalize_money_column(original_df["支出"])

    refund_df = _read_refund_list(refund_path)
    if refund_df.empty:
        log.warning("💡 退款清单为空，将原样复制账单")
        cleaned = original_df.copy()
    else:
        # 构造匹配 key：(交易日, 描述, 金额)
        refund_keys = set()
        for _, row in refund_df.iterrows():
            for side, amt_col in [("存入", "存入金额"), ("支出", "支出金额")]:
                pass
            for prefix, date_col, desc_col, amt_col in [
                ("存入", "存入交易日", "存入描述", "存入金额"),
                ("支出", "支出交易日", "支出描述", "支出金额"),
            ]:
                if date_col in row and desc_col in row and amt_col in row:
                    key = (
                        str(row[date_col]).strip(),
                        str(row[desc_col]).strip(),
                        parse_money(row[amt_col]),
                    )
                    refund_keys.add(key)

        log.info(f"   退款记录 {len(refund_keys)} 条")

        # 在原始账单中标记
        cleaned_rows = []
        for _, row in original_df.iterrows():
            trans_date = str(row["交易日"]).strip()
            trans_desc = str(row["交易描述"]).strip()
            for amt in (row["存入"], row["支出"]):
                if amt > 0 and (trans_date, trans_desc, amt) in refund_keys:
                    break
            else:
                cleaned_rows.append(row)
        cleaned = pd.DataFrame(cleaned_rows).reset_index(drop=True)

    log.info(f"   原始 {len(original_df)} 条 → 清理后 {len(cleaned)} 条")

    # 补齐摘要列
    for col in SUMMARY_COLUMNS:
        if col in original_df.columns and col not in cleaned.columns:
            cleaned[col] = original_df[col].iloc[0] if not original_df.empty else None

    # 退款清单（标准化）
    if not refund_df.empty:
        refund_out = refund_df[[
            "存入交易日", "存入描述", "存入金额",
            "支出交易日", "支出描述", "支出金额",
        ]].copy()
        refund_out.columns = [
            "存入交易日", "存入描述", "存入金额",
            "支出交易日", "支出描述", "支出金额",
        ]
    else:
        refund_out = pd.DataFrame(columns=[
            "存入交易日", "存入描述", "存入金额",
            "支出交易日", "支出描述", "支出金额",
        ])

    # 加合计行
    original_with_total = add_total_row(original_df[TRANSACTION_COLUMNS])
    refund_with_total = _add_refund_total_row(refund_out)
    cleaned_with_total = add_total_row(cleaned[TRANSACTION_COLUMNS])

    # 写 Excel
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        original_with_total.to_excel(writer, sheet_name="原始账单", index=False)
        refund_with_total.to_excel(writer, sheet_name="本期退款账单", index=False)
        cleaned_with_total.to_excel(writer, sheet_name="剔除退款后账单", index=False)

    wb = load_workbook(output_path)
    for name in ["原始账单", "本期退款账单", "剔除退款后账单"]:
        ws = wb[name]
        apply_header_format(ws)
        apply_body_border(ws)
        auto_column_width(ws)
        # 末行合计格式化
        if ws.max_row >= 2:
            last_row = ws.max_row
            first_cell = ws.cell(row=last_row, column=1).value
            if first_cell == "合计":
                apply_total_row_format(ws)

    # 给"剔除退款后账单"中 >阈值的支出标黄（用于人工介入时同步核验大额消费）
    if "剔除退款后账单" in wb.sheetnames:
        ws_cleaned = wb["剔除退款后账单"]
        expense_col_idx = TRANSACTION_COLUMNS.index("支出") + 1
        highlight_high_value_expenses(
            ws_cleaned, expense_col=expense_col_idx, threshold=HIGHLIGHT_THRESHOLD
        )
        log.info(f"   已对清理结果中 >{HIGHLIGHT_THRESHOLD} 元的支出标黄")

    wb.save(output_path)

    log.info(f"✅ 清理结果: {output_path}")
    return output_path


def _add_refund_total_row(df: pd.DataFrame) -> pd.DataFrame:
    """给退款清单添加合计行（金额列用公式）"""
    if df.empty:
        return df
    df = df.copy()
    n = len(df)
    df.loc[n] = {
        "存入交易日": "合计",
        "存入描述": "",
        "存入金额": f"=SUM(C2:C{n+1})",
        "支出交易日": "",
        "支出描述": "",
        "支出金额": f"=SUM(F2:F{n+1})",
    }
    return df
