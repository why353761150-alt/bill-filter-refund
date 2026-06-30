"""
账单统计分析 + 图表生成。

输入：data/working/<账期>/分析账单.xlsx
      （人工填"交易类别"列，参考下面的提示词）

DeepSeek 提示词（也可用于手工分类）：

    这是我的账单，我要对每笔消费进行分类，旅行，娱乐，网购，家居，通勤，
    通信费，物业，水费、电费、燃气费、你帮我按照以上分类对消费进行分类，
    如果拿不准的就空着，最后按照原表格的形式，我可以直接复制的方式输出给我。

    如果摘要中明确有"旅行"、"机票"、"酒店"，或者全是英文，等，归为旅行。
    如果摘要中明确有"电影"、"娱乐"、"酒吧"、"美团"等，归为娱乐。
    如果摘要中明确有"淘宝"、"京东"、"拼多多"等，归为网购。
    如果摘要中明确有"公交"、"地铁"、"巴士"、"打车"、"哈罗"等，归为通勤。
    如果摘要中明确有"手机充值"、"话费"等，归为通信费。
    如果摘要中明确有"物业"，归为物业。
    如果摘要中明确有"水费"、"电费"、"燃气费"，分别对应。"生活缴费"标注为待定。
"""
from __future__ import annotations

import re
from pathlib import Path
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils.dataframe import dataframe_to_rows

from .config import (
    ANALYSIS_COLUMNS,
    ANALYSIS_INPUT_NAME,
    ANALYSIS_REPORT_PREFIX,
    log,
    period_dirs,
    identify_period,
    current_period,
)
from .excel_utils import parse_money


# 支付方式识别
PAYMENT_KEYWORDS = {
    "支付宝": ["支付宝"],
    "微信": ["微信"],
    "京东": ["京东"],
    "财付通": ["财付通"],
    "云闪付": ["云闪付", "银联"],
}


def identify_payment(desc: str) -> str:
    if not isinstance(desc, str):
        return "其他"
    for name, kws in PAYMENT_KEYWORDS.items():
        if any(kw in desc for kw in kws):
            return name
    return "其他"


def analyze(analysis_input_path: Path | str) -> Path:
    """
    读取分析账单.xlsx，生成账单分析报告。

    修复原 nan 月份 bug：从路径中识别账期，而不是从数据中推断。
    """
    src = Path(analysis_input_path)
    if not src.exists():
        raise FileNotFoundError(
            f"找不到 {ANALYSIS_INPUT_NAME}: {src}\n"
            f"请先从最终账单复制支出列到 working 目录并填好'交易类别'"
        )

    period = identify_period(str(src)) or current_period()
    dirs = period_dirs(period)
    output_path = dirs["output"] / f"{ANALYSIS_REPORT_PREFIX}{period}.xlsx"

    log.info(f"📊 分析账单: {src.name}")
    log.info(f"📅 账期: {period}")

    df = pd.read_excel(src)
    # 兼容中文/英文列名
    col_map = {}
    for c in df.columns:
        c_str = str(c).strip()
        if "交易金额" in c_str or "金额" in c_str:
            col_map[c] = "交易金额（RMB）"
        elif c_str in ("交易日", "日期"):
            col_map[c] = "交易日"
        elif "摘要" in c_str or "描述" in c_str:
            col_map[c] = "交易摘要"
        elif "类别" in c_str:
            col_map[c] = "交易类别"
    df = df.rename(columns=col_map)

    # 确保有必需的列
    for col in ANALYSIS_COLUMNS:
        if col not in df.columns:
            log.error(f"分析账单缺少列: {col}")
            log.error(f"实际列: {list(df.columns)}")
            raise ValueError(f"分析账单缺少必需列: {col}")

    df["交易金额（RMB）"] = df["交易金额（RMB）"].apply(parse_money)
    df["交易日"] = pd.to_datetime(df["交易日"], errors="coerce")
    df = df.dropna(subset=["交易日", "交易金额（RMB）"])
    df = df[df["交易金额（RMB）"] > 0].copy()

    if df.empty:
        log.warning("💡 没有有效的消费记录")
        return output_path

    total = df["交易金额（RMB）"].sum()
    log.info(f"   共 {len(df)} 条记录，总支出 ¥{total:.2f}")

    # 中文字体
    plt.rcParams["font.sans-serif"] = ["SimHei", "Arial Unicode MS", "DejaVu Sans"]
    plt.rcParams["axes.unicode_minus"] = False

    wb = Workbook()
    wb.remove(wb.active)

    # ========== Sheet 1: 月度汇总 ==========
    ws = wb.create_sheet("月度汇总")
    daily = df.groupby(df["交易日"].dt.date)["交易金额（RMB）"].sum()
    max_day = daily.idxmax() if not daily.empty else None
    min_day = daily.idxmin() if not daily.empty else None
    summary = [
        ["项目", "数值", "说明"],
        ["账期", period, "账单所属月份"],
        ["总支出", f"¥{total:.2f}", "本月所有支出总和"],
        ["总交易笔数", len(df), "本月交易总笔数"],
        ["消费天数", len(daily), "有消费记录的天数"],
        ["日均消费", f"¥{daily.mean():.2f}" if not daily.empty else "¥0.00", "平均每天消费"],
        ["最高单笔", f"¥{df['交易金额（RMB）'].max():.2f}", "最高单笔消费"],
        ["平均单笔", f"¥{df['交易金额（RMB）'].mean():.2f}", "平均每笔消费"],
        ["最高日消费", f"{max_day} ¥{daily[max_day]:.2f}" if max_day else "-", "消费最高日"],
        ["最低日消费", f"{min_day} ¥{daily[min_day]:.2f}" if min_day else "-", "消费最低日"],
    ]
    for r, row in enumerate(summary, start=1):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = Font(bold=(r == 1), size=11)
            if r == 1:
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            cell.alignment = Alignment(horizontal="center" if c != 3 else "left",
                                       vertical="center")
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 30

    # ========== Sheet 2: 支出类别统计 ==========
    cat_stats = df.groupby("交易类别")["交易金额（RMB）"].agg(
        总金额="sum", 笔数="count", 均值="mean"
    ).round(2).sort_values("总金额", ascending=False).reset_index()
    cat_stats["占比(%)"] = (cat_stats["总金额"] / total * 100).round(2)

    # 总计行
    total_row = pd.DataFrame([{
        "交易类别": "总计",
        "总金额": total,
        "笔数": len(df),
        "均值": df["交易金额（RMB）"].mean().round(2),
        "占比(%)": 100.0,
    }])
    cat_table = pd.concat([cat_stats, total_row], ignore_index=True)

    ws_cat = wb.create_sheet("支出类别")
    for c, header in enumerate(["消费类别", "总金额", "交易笔数", "平均金额", "占比(%)"], start=1):
        cell = ws_cat.cell(row=1, column=c, value=header)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    for r, row in enumerate(dataframe_to_rows(cat_table, index=False, header=False), start=2):
        for c, val in enumerate(row, start=1):
            cell = ws_cat.cell(row=r, column=c, value=val)
            cell.font = Font(bold=(row[0] == "总计"), size=11)
            if row[0] == "总计":
                cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    for col, w in zip("ABCDE", [18, 15, 12, 15, 12]):
        ws_cat.column_dimensions[col].width = w

    if not cat_stats.empty:
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = f"{period} 支出类别"
        chart.x_axis.title = "金额（元）"
        chart.y_axis.title = "类别"
        data = Reference(ws_cat, min_col=2, min_row=1,
                         max_row=1 + len(cat_stats))
        cats = Reference(ws_cat, min_col=1, min_row=2,
                         max_row=1 + len(cat_stats))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width = 22
        chart.height = 15
        ws_cat.add_chart(chart, "G2")

    # ========== Sheet 3: 每日消费 ==========
    daily_df = daily.reset_index()
    daily_df.columns = ["日期", "消费金额"]
    daily_df = daily_df.sort_values("日期")
    daily_df["排名"] = daily_df["消费金额"].rank(ascending=False, method="min").astype(int)
    daily_df["累计"] = daily_df["消费金额"].cumsum()
    daily_df["累计占比(%)"] = (daily_df["累计"] / total * 100).round(2)

    ws_day = wb.create_sheet("每日消费")
    for c, header in enumerate(["日期", "消费金额", "排名", "累计", "累计占比(%)"], start=1):
        cell = ws_day.cell(row=1, column=c, value=header)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    for r, row in enumerate(dataframe_to_rows(daily_df, index=False, header=False), start=2):
        for c, val in enumerate(row, start=1):
            cell = ws_day.cell(row=r, column=c, value=val)
            cell.font = Font(size=11)
            if c == 5:
                cell.number_format = "0.00"
    for col, w in zip("ABCDE", [14, 14, 8, 14, 14]):
        ws_day.column_dimensions[col].width = w

    if not daily_df.empty:
        chart2 = BarChart()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = f"{period} 每日消费"
        chart2.x_axis.title = "日期"
        chart2.y_axis.title = "金额（元）"
        data2 = Reference(ws_day, min_col=2, min_row=1, max_row=1 + len(daily_df))
        chart2.add_data(data2, titles_from_data=True)
        chart2.width = 30
        chart2.height = 12
        ws_day.add_chart(chart2, "G2")

    # ========== Sheet 4: 支付方式统计 ==========
    df["支付方式"] = df["交易摘要"].apply(identify_payment)
    pay_stats = df.groupby("支付方式")["交易金额（RMB）"].agg(
        总金额="sum", 笔数="count"
    ).round(2)
    pay_stats["占比(%)"] = (pay_stats["总金额"] / total * 100).round(2)
    pay_stats = pay_stats.sort_values("总金额", ascending=False).reset_index()

    ws_pay = wb.create_sheet("支付方式")
    for c, header in enumerate(["支付方式", "总金额", "笔数", "占比(%)"], start=1):
        cell = ws_pay.cell(row=1, column=c, value=header)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    for r, row in enumerate(dataframe_to_rows(pay_stats, index=False, header=False), start=2):
        for c, val in enumerate(row, start=1):
            cell = ws_pay.cell(row=r, column=c, value=val)
            cell.font = Font(size=11)
    for col, w in zip("ABCD", [14, 14, 10, 12]):
        ws_pay.column_dimensions[col].width = w

    if not pay_stats.empty:
        pie = PieChart()
        pie.title = f"{period} 支付方式占比"
        data3 = Reference(ws_pay, min_col=2, min_row=1, max_row=1 + len(pay_stats))
        labels = Reference(ws_pay, min_col=1, min_row=2, max_row=1 + len(pay_stats))
        pie.add_data(data3, titles_from_data=True)
        pie.set_categories(labels)
        pie.dataLabels = DataLabelList(showPercent=True, showCatName=True)
        pie.width = 18
        pie.height = 12
        ws_pay.add_chart(pie, "F2")

    wb.save(output_path)
    log.info(f"✅ 分析报告: {output_path}")
    return output_path
